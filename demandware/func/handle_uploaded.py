import json
from django.conf import settings
from openpyxl import load_workbook
from openpyxl.utils import cell as pycell
from django.apps import apps
from demandware.models import ProductMaster, ProductMeta, RelatedProduct, Category, CategoryMeta, Variant, ProductImage, HeaderMgr
# Imports the Google Cloud client library
from google.cloud import translate
import logging

# Instantiates a client
translate_client = translate.Client()

# Get an instance of a logger
logger = logging.getLogger('django')

def handle_uploaded_file(form=None, f=None, model_name=None):
	wb = load_workbook(f)
	print(wb.sheetnames)
	# ws = wb.active
	# ws=wb.get_sheet_by_name("data")
	sheet_name = form.cleaned_data.get('sheet_name')
	data_type = form.cleaned_data.get('data_type')
	min_row = form.cleaned_data.get('from_row')
	max_row = form.cleaned_data.get('to_row')
	min_col = form.cleaned_data.get('from_col')
	max_col = form.cleaned_data.get('to_col')
	header_row = form.cleaned_data.get('header_row')
	truncate = form.cleaned_data.get('truncate')
	
	model = apps.get_model(app_label='demandware', model_name=data_type)
	if sheet_name == None or sheet_name == '':
		sheet_name = 'data'
	ws = wb[sheet_name]

	if max_row == 0:
		max_row = ws.max_row

	if max_col == '':
		max_col = ws.max_column

	min_col_num = pycell.column_index_from_string(min_col)
	max_col_num = pycell.column_index_from_string(max_col)
	dataset = ws.iter_rows(min_row=min_row, min_col=min_col_num, max_col=max_col_num, max_row=max_row)

	header = ws.iter_rows(min_row=header_row, min_col=min_col_num, max_col=max_col_num, max_row=header_row)
	header = [cell.value for cell in list(header)[0]]

	_count = 0
	insertDataSet = []
	extInsertDataSet = []
	fields = get_model_fields(model)
	print(fields)
	for row in dataset:
		values = {}
		extValues = {}
		for key, _cell in zip(header, row):
			if key in fields:
				values[key] = _cell.value if _cell.value else ''
			else:
				extValues[key] = _cell.value if _cell.value else ''
		_count = _count + 1
		insertDataSet.append(values)
		extInsertDataSet.append(extValues)

	result = detect_service(model_name=data_type, data=insertDataSet, extData=extInsertDataSet, header=header, truncate=truncate)
	return result

def detect_service(model_name=None, data=None, extData=None, header=None, truncate=False):
	if model_name == 'productmaster':
		if truncate:
			truncate_table('dtb_product_master_extra');
			truncate_table('dtb_product_master');
		return product_master_process(data=data, extData=extData, header=header)

	if model_name == 'relatedproduct':
		if truncate:
			truncate_table('dtb_related_product');
		return related_product_process(data=data, extData=extData)

	if model_name == 'variant':
		if truncate:
			truncate_table('dtb_variant');
		return variant_process(data=data, extData=extData)

	if model_name == 'productimage':
		if truncate:
			truncate_table('dtb_product_image');
		return product_image_process(data=data, extData=extData)

	if model_name == 'category':
		if truncate:
			truncate_table('dtb_category_extra');
			truncate_table('dtb_categories');
		return category_process(data=data, header=header, extData=extData)

	if model_name == 'productcategory':
		if truncate:
			truncate_table('dtb_product_category');
		return product_category_process(data=data, extData=extData)

	return None


def get_model_fields(model):
    return [f.name for f in model._meta.fields]


def product_master_process(data=None, header=None, extData=None):
	import re
	from collections import OrderedDict
	"""
	1. Insert header to header_mgr db
	2. If header name is existing in table, then insert to table
		Else insert or update to metadata
	"""
	from demandware.models_handler.header_handler import insert_bulk_header
	from demandware.models_handler.product_handler import insert_product_master, insert_bulk_product_master, insert_product_metadata
	result = None

	### Insert header data
	# result = insert_bulk_header(header_list=header, header_type=HeaderMgr.PRODUCT)

	### Insert data
	### 前処理
	# product_commentaryを処理する | product_all_colorを処理する
	
	for idx, items in zip(range(len(extData)), extData):
		if data[idx]['product_id'] == None or str(data[idx]['product_id']).strip() == '':
			continue
		commentary = {}
		all_color = {}
		_items = list(items.items())
		for key, value in _items:
			_re_commentary = re.compile("product_commentary_(.*)_(.*)")
			_search_commentary = _re_commentary.search(key)
			if _search_commentary != None:
				commentary[_search_commentary.group(1)] = commentary[_search_commentary.group(1)] if _search_commentary.group(1) in commentary else {}
				commentary[_search_commentary.group(1)][_search_commentary.group(0)] = value
				del extData[idx][key]

			_re_all_color = re.compile("product_all_color_(.*)_(\d+)")
			_search_all_color = _re_all_color.search(key)
			if _search_all_color != None:
				all_color[_search_all_color.group(2)] = all_color[_search_all_color.group(2)] if _search_all_color.group(2) in all_color else {}
				color_format = value
				if _search_all_color.group(1) == 'hexa_code' and value != None and value != "":
					if len(str(value).split(",")) == 3:
						color_format = 'rgb(' + str(value) + ')'
					elif len(str(value).split(",")) > 1:
						print("[ERROR] FORMAT OF COLOR IS INCORRECT!" + color_format)

				all_color[_search_all_color.group(2)][_search_all_color.group(0)] = color_format
				del extData[idx][key]

		commentary = sorted(commentary.items(), key=lambda t: int(t[0]))
		commentary = [item[1] for item in commentary]
		all_color = sorted(all_color.items(), key=lambda t: int(t[0]))
		all_color = [item[1] for item in all_color]
		data[idx]['product_commentary'] = json.dumps(commentary, ensure_ascii=False)
		data[idx]['product_all_color'] = json.dumps(all_color, ensure_ascii=False)

	result = insert_bulk_product_master(data=data)
	# result = insert_product_metadata(data=data, extData=extData)
	return result

def related_product_process(data=None, extData=None):
	from demandware.models_handler.product_handler import insert_related_product
	return insert_related_product(data)

def variant_process(data=None, extData=None):
	from demandware.models_handler.product_handler import insert_variant
	return insert_variant(data)

def product_image_process(data=None, extData=None):
	from demandware.models_handler.product_handler import insert_product_image
	return insert_product_image(data)

def category_process(data=None, header=None, extData=None):
	from demandware.models_handler.header_handler import insert_bulk_header
	from demandware.models_handler.category_handler import insert_category
	
	result = dict(
		message=[],
		count=0
	)

	### Insert header data
	insert_bulk_header(header_list=header, header_type=HeaderMgr.CATEGORY)

	### Insert data
	countLevel1=0
	countLevel2=0
	countLevel3=0
	categoryLevel1ToReset=''
	categoryLevel2ToReset=''
	categoryLevel3ToReset=''
	index=0
	for item in extData:
		if item['category_level_3_name'] == 'すべて':
			continue;
		if item['category_level_1_id'] == None or item['category_level_1_id'] == '':
			continue
		if(categoryLevel1ToReset == item['category_level_1_id']):
			if(categoryLevel2ToReset == item['category_level_2_id']):
				countLevel3 = countLevel3 + 1
			else:
				countLevel2 = countLevel2 + 1
				categoryLevel2ToReset = item['category_level_2_id']
				countLevel3 = 0
				categoryLevel3ToReset = item['category_level_3_id']
		else:
			countLevel3 = 0
			categoryLevel3ToReset = item['category_level_3_id']
			countLevel2 = 0
			categoryLevel2ToReset = item['category_level_2_id']
			countLevel1 = countLevel1 + 1
			categoryLevel1ToReset = item['category_level_1_id']

		datalv1 = dict(
			category_id=item['category_level_1_id'],
			category_name=item['category_level_1_name'],
			category_level=1,
			category_parent=None,
			category_custom_url=item['category_level_1_id'],
			category_position=countLevel1,
		)
		if 'category_level_1_name_JP' in item:
			datalv1['category_name_jp']=item['category_level_1_name_JP']

		print("Insert category: " + str(datalv1['category_id']))
		result1 = insert_category(data=datalv1)
		result['count'] = result['count'] + 1;
		if result1['error']:
			result['message'].append(result1['error'])
			return result
		if item['category_level_2_id'] == None or item['category_level_2_id'] == '':
			continue

		datalv2 = dict(
			category_id="%s-%s" % (result1['obj'].category_id, item['category_level_2_id']),
			category_name=item['category_level_2_name'],
			category_level=2,
			category_parent=result1['obj'],
			category_custom_url=item['category_level_2_id'],
			category_position=countLevel2,
		)
		if 'category_level_2_name_JP' in item:
			datalv1['category_name_jp']=item['category_level_2_name_JP']

		print("Insert category: " + str(datalv2['category_id']))
		result2 = insert_category(data=datalv2)
		result['count'] = result['count'] + 1;
		if result2['error']:
			result['message'].append(result2['error'])
			return result
		if item['category_level_3_id'] == None or item['category_level_3_id'] == '':
			continue

		datalv3 = dict(
			category_id="%s-%s" % (result2['obj'].category_id, item['category_level_3_id']),
			category_name=item['category_level_3_name'],
			category_name_jp=item['category_level_3_name'],
			category_level=3,
			category_parent=result2['obj'],
			category_custom_url=item['category_level_3_id'],
			category_position=countLevel3,
		)
		
		print("Insert category: " + str(datalv3['category_id']))
		result3 = insert_category(data=datalv3)
		result['count'] = result['count'] + 1;
		if result3['error']:
			result['message'].append(result3['error'])
			return result
		index=index+1

	print('Done!')
	return result

def product_category_process(data=None, extData=None):
	from demandware.models_handler.category_handler import insert_product_category
	return insert_product_category(data)

def truncate_table(db_table=None):
	from django.db import connection
	cursor = connection.cursor()
	cursor.execute('SET FOREIGN_KEY_CHECKS = 0; TRUNCATE TABLE `{0}`; SET FOREIGN_KEY_CHECKS = 1;'.format(db_table))
